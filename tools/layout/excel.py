"""Render the 島配置図 workbook with openpyxl.

Sheets produced:

* one **detail** sheet per hall (``東4_詳細`` …): 1 cell = 1 circle (number + side),
  islands coloured per block, walls highlighted, and the 中央通路 / 細通路 cross-aisles
  drawn as shaded bands. Within a band the islands are aligned by their 細通路 (per the
  spec) and serpentine numbers run bottom-up (右下 = bottom of the column).
* an **島一覧** schematic sheet: 1 cell = 1 island (prefix + 番号数);
* a **distances** sheet: the cluster×cluster matrix (matches ``hall_distances.csv``);
* a **blocks** sheet: a flat table mirror of ``block_layouts.csv``.
"""

from __future__ import annotations

import math
import re
from pathlib import Path

from openpyxl import Workbook
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from .model import (
    BLOCK_CSV_HEADER,
    BlockRecord,
    EventConfig,
    Island,
    Row,
    corridor_geometry,
    expand_island_local,
    hall_distance,
)

# --- styling ----------------------------------------------------------------

_PALETTE = [
    "FFF1F0", "E8F1FF", "E9F7E9", "FFF6E0", "F3E8FF", "E0F7F7",
    "FCE8F0", "EEF3D8", "E8ECF8", "F7EAD9", "E5F0E0", "F0E5F0",
]
_WALL_COLOR = "FFD966"
_HEADER_COLOR = "D9D9D9"
_MAJOR_COLOR = "F4B183"  # 太通路 (中央通路 / 大通路)
_MINOR_COLOR = "DDDDDD"  # 細通路 / 横通路 / ブロック境界

_BOLD = Font(bold=True)
_SMALL = Font(size=8)
_TINY = Font(size=7, color="808080")
_CENTER = Alignment(horizontal="center", vertical="center")
_WRAP = Alignment(horizontal="center", vertical="center", wrap_text=True)
_THIN = Side(style="thin", color="BBBBBB")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)


def _fill(color: str) -> PatternFill:
    return PatternFill(start_color=color, end_color=color, fill_type="solid")


def _safe_title(name: str, used: set[str]) -> str:
    """A unique, ≤31-char sheet title with the characters Excel forbids stripped."""
    title = re.sub(r"[\[\]:*?/\\]", "", name)[:31] or "sheet"
    base, i = title, 1
    while title in used:
        suffix = f"_{i}"
        title = base[: 31 - len(suffix)] + suffix
        i += 1
    used.add(title)
    return title


class _ColorBook:
    """Hands out a stable fill per (hall, prefix), cycling the palette."""

    def __init__(self) -> None:
        self._map: dict[tuple[str, str], str] = {}

    def fill_for(self, hall_label: str, island: Island) -> PatternFill:
        if island.kind == "wall":
            return _fill(_WALL_COLOR)
        key = (hall_label, island.prefix)
        if key not in self._map:
            self._map[key] = _PALETTE[len(self._map) % len(_PALETTE)]
        return _fill(self._map[key])


def _col_offset(face: int, side: str) -> int:
    """Column within an island group: face0(a,b) = 0,1 ; face1(a,b) = 2,3."""
    return face * 2 + (0 if side == "a" else 1)


# --- per-island vertical layout (cells + corridor gap bands, bottom-up) ------


class _IslandLayout:
    """Where each circle and corridor band sits in an island's own column stack.

    Visual rows are counted from the bottom (``0`` = the near/右下 end). A corridor
    inserts a blank band (2 rows for a 太通路, 1 for a 細通路) that pushes the circles
    above it up. ``ref_vis`` marks the lowest 細通路 — the row a whole band of islands is
    aligned on, so their thin corridors line up across the sheet.
    """

    def __init__(self, island: Island, pitch: float) -> None:
        cells = expand_island_local(island)
        cors = [corridor_geometry(island, c, pitch) for c in island.corridors]

        gaps: dict[int, tuple[int, str]] = {}  # after-depth -> (thickness, kind)
        for g in cors:
            after = int(math.floor(g["cross_along"] / pitch + 0.5))
            thick = 2 if g["kind"] == "major" else 1
            if after not in gaps or thick > gaps[after][0]:
                gaps[after] = (thick, g["kind"])

        def cumul(depth: int) -> int:
            return sum(th for ad, (th, _) in gaps.items() if ad < depth)

        self.cells = cells
        self._gaps = gaps
        self._cumul = cumul
        max_depth = max(c.depth for c in cells)
        self.total_height = (max_depth - 1) + cumul(max_depth + 1) + 1
        self.bands = [(self.vis(ad) + 1, th, kind) for ad, (th, kind) in gaps.items()]

        minor = [g["cross_along"] for g in cors if g["kind"] == "minor"]
        if minor:
            ref_after = int(math.floor(min(minor) / pitch + 0.5))
            self.ref_vis = self.vis(ref_after) + 1
        else:
            self.ref_vis = 0

    def vis(self, depth: int) -> int:
        """Visual row (from the bottom) of a circle at this depth."""
        return (depth - 1) + self._cumul(depth)


# --- detail sheet -----------------------------------------------------------


def _render_row(ws: Worksheet, top: int, row: Row, hall_label: str, colors: _ColorBook,
                pitch: float) -> int:
    """Render one band/wall run starting at sheet row ``top``; return the next free row."""
    ws.cell(top, 1, row.label or hall_label).font = _BOLD
    label_row = top + 1
    content_top = top + 2

    layouts = [_IslandLayout(isl, pitch) for isl in row.islands]
    max_above = max((lay.total_height - 1 - lay.ref_vis for lay in layouts), default=0)
    max_below = max((lay.ref_vis for lay in layouts), default=0)
    ref_row = content_top + max_above  # sheet row the 細通路 reference line sits on

    col = 1
    for island, lay in zip(row.islands, layouts):
        group_cols = 2 if island.kind == "wall" else 4
        fill = colors.fill_for(hall_label, island)

        label = ws.cell(label_row, col, island.prefix)
        label.font = _BOLD
        label.alignment = _CENTER
        if group_cols > 1:
            ws.merge_cells(start_row=label_row, start_column=col,
                           end_row=label_row, end_column=col + group_cols - 1)

        # Circles: visual row v → sheet row ref_row - (v - ref_vis).
        for lc in lay.cells:
            v = lay.vis(lc.depth)
            r = ref_row - (v - lay.ref_vis)
            cell = ws.cell(r, col + _col_offset(lc.face, lc.side), lc.number)
            cell.fill = fill
            cell.font = _SMALL
            cell.alignment = _CENTER
            cell.border = _BORDER

        # Corridor bands across the island's columns.
        for start_v, thick, kind in lay.bands:
            band_fill = _fill(_MAJOR_COLOR if kind == "major" else _MINOR_COLOR)
            for t in range(thick):
                r = ref_row - (start_v + t - lay.ref_vis)
                for dc in range(group_cols):
                    bc = ws.cell(r, col + dc)
                    bc.fill = band_fill
                    bc.border = _BORDER
            if kind == "major":
                tag = ws.cell(ref_row - (start_v - lay.ref_vis), col, "中央")
                tag.font = _TINY
                tag.alignment = _CENTER

        col += group_cols + 1

    return ref_row + max_below + 2


def _build_detail_sheet(wb: Workbook, hall, colors: _ColorBook, used: set[str],
                        pitch: float) -> None:
    ws = wb.create_sheet(_safe_title(f"{hall.label}_詳細", used))
    top = 1
    max_col = 1
    for row in hall.rows:
        top = _render_row(ws, top, row, hall.label, colors, pitch)
        cols = sum((2 if isl.kind == "wall" else 4) + 1 for isl in row.islands)
        max_col = max(max_col, cols)
    for c in range(1, max_col + 2):
        ws.column_dimensions[get_column_letter(c)].width = 3.6
    ws.freeze_panes = "A1"
    ws.sheet_view.zoomScale = 80


# --- island schematic sheet -------------------------------------------------


def _build_overview_sheet(wb: Workbook, event: EventConfig, colors: _ColorBook) -> None:
    ws = wb.create_sheet("島一覧")
    ws.cell(1, 1, f"{event.event} 島一覧 (1セル=1島)").font = Font(bold=True, size=12)
    r = 3
    for hall in event.halls:
        head = ws.cell(r, 1, f"{hall.label}  [{hall.cluster}]")
        head.font = _BOLD
        head.fill = _fill(_HEADER_COLOR)
        r += 1
        for row in hall.rows:
            ws.cell(r, 1, row.label).font = _SMALL
            c = 2
            for island in row.islands:
                cell = ws.cell(r, c, f"{island.prefix}\n{island.n_max}")
                cell.fill = colors.fill_for(hall.label, island)
                cell.alignment = _WRAP
                cell.border = _BORDER
                c += 1
            r += 1
        r += 1
    ws.column_dimensions["A"].width = 14
    for c in range(2, 40):
        ws.column_dimensions[get_column_letter(c)].width = 6
    ws.freeze_panes = "B1"


# --- distances sheet --------------------------------------------------------


def _build_distances_sheet(wb: Workbook, event: EventConfig) -> None:
    ws = wb.create_sheet("distances")
    ids = [c.id for c in event.clusters]
    ws.cell(1, 1, "cluster").font = _BOLD
    for j, cid in enumerate(ids):
        ws.cell(1, 2 + j, cid).font = _BOLD
        ws.cell(2 + j, 1, cid).font = _BOLD
    for i, a in enumerate(ids):
        for j, b in enumerate(ids):
            d = hall_distance(event.hall_distances, a, b)
            cell = ws.cell(2 + i, 2 + j)
            cell.alignment = _CENTER
            cell.border = _BORDER
            if d is not None:
                cell.value = d

    first, last = "B2", f"{get_column_letter(1 + len(ids))}{1 + len(ids)}"
    ws.conditional_formatting.add(
        f"{first}:{last}",
        ColorScaleRule(
            start_type="min", start_color="F8FCE8",
            mid_type="percentile", mid_value=50, mid_color="FFE08A",
            end_type="max", end_color="F4978E",
        ),
    )
    legend_row = len(ids) + 4
    ws.cell(legend_row, 1, "凡例 (cluster → ホール)").font = _BOLD
    for k, c in enumerate(event.clusters):
        ws.cell(legend_row + 1 + k, 1, c.id)
        ws.cell(legend_row + 1 + k, 2, c.label)
    ws.column_dimensions["A"].width = 10
    for c in range(2, 2 + len(ids)):
        ws.column_dimensions[get_column_letter(c)].width = 9


# --- blocks table sheet -----------------------------------------------------


def _build_blocks_sheet(wb: Workbook, records: list[BlockRecord]) -> None:
    ws = wb.create_sheet("blocks")
    for j, name in enumerate(BLOCK_CSV_HEADER, start=1):
        cell = ws.cell(1, j, name)
        cell.font = _BOLD
        cell.fill = _fill(_HEADER_COLOR)
    for i, rec in enumerate(records, start=2):
        values = [
            rec.id, rec.building, rec.hall, rec.anchor_x, rec.anchor_y, rec.along,
            rec.cross, rec.n_max, rec.face0_len, rec.pitch, rec.island_width, rec.kind,
            rec.cluster, rec.along_deg, rec.cross_deg,
            ";".join(f"{along}@{kind}" for along, kind in rec.crossings),
            rec.number_base,
        ]
        for j, value in enumerate(values, start=1):
            ws.cell(i, j, value)
    ws.freeze_panes = "A2"
    widths = [10, 7, 5, 9, 9, 6, 6, 6, 8, 6, 11, 7, 7, 9, 9, 16, 8]
    for j, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(j)].width = w


# --- public entry -----------------------------------------------------------


def build_workbook(event: EventConfig, records: list[BlockRecord], path: Path) -> None:
    """Write the full ``layout_<event>.xlsx`` workbook to ``path``."""
    wb = Workbook()
    wb.remove(wb.active)  # drop the default empty sheet
    colors = _ColorBook()
    used: set[str] = set()

    for hall in event.halls:
        _build_detail_sheet(wb, hall, colors, used, event.pitch_m)
    _build_overview_sheet(wb, event, colors)
    _build_distances_sheet(wb, event)
    _build_blocks_sheet(wb, records)

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
